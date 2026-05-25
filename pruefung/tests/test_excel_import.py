"""Tests für pruefung.excel_import + /api/import-excel-schablone.

Deterministischer Import — kein LLM involviert, deshalb können wir hier
End-zu-End mit echten XLSX-Bytes testen.
"""
from __future__ import annotations

from io import BytesIO
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient

from pruefung.excel_import import (
    SCHABLONE_HELFERLISTE_HEADERS,
    import_helferliste,
)
from pruefung.main import app


client = TestClient(app)


def _baue_xlsx(headers: list, daten: list[tuple]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row in enumerate(daten, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_import_helferliste_basis():
    xlsx = _baue_xlsx(
        SCHABLONE_HELFERLISTE_HEADERS,
        [
            (1, "Müller", "Anna", "Wohnbereich Süd", "2018-03-01", None, 12.0, 144.0),
            (2, "Schmidt", "Klaus", "Begleitdienst", None, None, 8.0, 96.0),
        ],
    )
    result = import_helferliste(xlsx)
    assert len(result) == 2
    assert result[0]["name"] == "Müller"
    assert result[0]["stunden_jahr"] == 144.0
    assert result[1]["eintritt"] is None


def test_import_helferliste_filtert_unvollstaendige():
    """Zeilen ohne Name oder Vorname werden übersprungen."""
    xlsx = _baue_xlsx(
        SCHABLONE_HELFERLISTE_HEADERS,
        [
            (1, "Müller", "Anna", None, None, None, None, None),
            (2, None,     "Anna", None, None, None, None, None),  # kein Name
            (3, "Schmidt", None,  None, None, None, None, None),  # kein Vorname
        ],
    )
    result = import_helferliste(xlsx)
    assert len(result) == 1
    assert result[0]["name"] == "Müller"


def test_import_helferliste_header_mismatch():
    """Falscher Header → ValueError mit Diagnostik."""
    xlsx = _baue_xlsx(
        ["Irgendwas", "anderes"],
        [("a", "b")],
    )
    with pytest.raises(ValueError, match="Schablonen-Mismatch"):
        import_helferliste(xlsx)


def test_import_helferliste_leeres_workbook():
    """openpyxl liefert für ein neues Workbook eine leere Zeile zurück —
    der Default-Sheet hat einen impliziten 'None'-Header. Unser Import
    behandelt das als Mismatch (keine Daten + falsche Header)."""
    wb = openpyxl.Workbook()
    buf = BytesIO()
    wb.save(buf)
    # Leere Workbook hat 1 leere Zeile → wird als Mismatch erkannt
    # ODER als 'rows is empty' behandelt. Beide Pfade sind ok für uns.
    try:
        result = import_helferliste(buf.getvalue())
        assert result == []
    except ValueError:
        pass


def test_import_helferliste_nur_header():
    """Header da, keine Daten → leere Liste."""
    xlsx = _baue_xlsx(SCHABLONE_HELFERLISTE_HEADERS, [])
    assert import_helferliste(xlsx) == []


# ── Endpoint ────────────────────────────────────────────────────────


def test_endpoint_import_excel_schablone_happy_path():
    xlsx = _baue_xlsx(
        SCHABLONE_HELFERLISTE_HEADERS,
        [(1, "Müller", "Anna", "X", None, None, 10.0, 120.0)],
    )
    r = client.post(
        "/api/import-excel-schablone",
        files={"file": ("h.xlsx", BytesIO(xlsx),
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["anzahl"] == 1
    assert body["helfer"][0]["name"] == "Müller"


def test_endpoint_import_excel_schablone_bad_header():
    xlsx = _baue_xlsx(["X", "Y"], [("a", "b")])
    r = client.post(
        "/api/import-excel-schablone",
        files={"file": ("h.xlsx", BytesIO(xlsx), "application/vnd.ms-excel")},
    )
    assert r.status_code == 400
    assert "Schablonen-Mismatch" in r.json()["detail"]


# ── Schablone selbst (smoke-Test) ───────────────────────────────────


_SCHABLONE = Path(__file__).resolve().parent.parent.parent / (
    "materialien/wuerzburg-2026/schablonen/helferliste.xlsx"
)


@pytest.mark.skipif(not _SCHABLONE.exists(),
                    reason="Schablone nicht generiert — scripts/erstelle_excel_schablonen.py")
def test_generierte_schablone_passt_zum_import():
    """Wenn jemand die Schablone generiert hat, MUSS sie ohne Fehler
    durch unseren eigenen Import gehen (sonst widersprechen sich die
    beiden Pfade)."""
    result = import_helferliste(_SCHABLONE.read_bytes())
    assert len(result) >= 1
    assert all("name" in r and "vorname" in r for r in result)
