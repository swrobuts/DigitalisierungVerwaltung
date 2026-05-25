"""Erzeugt die Excel-Schablonen, die /api/import-excel-schablone akzeptiert.

Aufruf manuell:
    uv run --project pruefung python scripts/erstelle_excel_schablonen.py

Die generierten Dateien landen in materialien/wuerzburg-2026/schablonen/
und werden vom Frontend zum Download angeboten (UE1 'Helferliste
hochladen' → Link 'Vorlage herunterladen').
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from pruefung.excel_import import SCHABLONE_HELFERLISTE_HEADERS


REPO_ROOT = Path(__file__).resolve().parent.parent
TARGET_DIR = REPO_ROOT / "materialien" / "wuerzburg-2026" / "schablonen"


def baue_helferliste_schablone() -> Path:
    """Erzeugt helferliste.xlsx mit Header + 3 Beispielzeilen."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Helferliste FB II"

    # Header
    header_fill = PatternFill("solid", fgColor="AD0E36")  # Würzburg-Rot
    header_font = Font(bold=True, color="FFFFFF")
    for col, name in enumerate(SCHABLONE_HELFERLISTE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    beispiele = [
        (1, "Müller", "Anna",  "Wohnbereich Süd",  "2018-03-01", None, 12.0, 144.0),
        (2, "Schmidt", "Klaus", "Begleitdienst",    "2020-09-15", None, 8.0,  96.0),
        (3, "Wagner", "Petra", "Cafeteria",        "2022-01-10", None, 16.0, 192.0),
    ]
    for row_idx, row in enumerate(beispiele, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Sinnvolle Spaltenbreiten
    for col_idx, width in enumerate([10, 18, 18, 25, 14, 14, 14, 14], start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    TARGET_DIR.mkdir(parents=True, exist_ok=True)
    out = TARGET_DIR / "helferliste.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    p = baue_helferliste_schablone()
    print(f"OK — Schablone geschrieben: {p}")
